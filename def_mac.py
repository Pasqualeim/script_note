import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import re
import tkinter as tk
from tkinter import filedialog

def select_file_component(prompt):
    """
    Mostra una finestra di dialogo per selezionare un file e restituisce il percorso.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    return file_path

def select_file_notes(prompt):
    """
    Mostra una finestra di dialogo per selezionare un file e restituisce il percorso.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    return file_path

def select_save_location(default_name):
    """
    Mostra una finestra di dialogo per selezionare la posizione di salvataggio del file.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        title="Seleziona dove salvare il file",
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    return file_path

def convert_version_format(version):
    """
    Converte una versione con punto (es. 7.22) in un numero intero (es. 722) per l'analisi.
    """
    try:
        return int(str(version).replace(".", "")) if pd.notna(version) else None
    except ValueError:
        return None

def extract_sp_level(sp_value):
    """
    Estrae il valore numerico da una stringa con prefisso 'SP' o da un numero con zeri iniziali.
    Esempio: 'SP007' -> 7, '0007' -> 7
    """
    if pd.isna(sp_value):
        return None
    sp_value_str = str(sp_value).strip()
    match = re.search(r"SP(\d+)", sp_value_str)
    if match:
        return int(match.group(1))
    elif sp_value_str.isdigit():
        return int(sp_value_str.lstrip("0")) if sp_value_str.lstrip("0") else 0
    return None

def check_release_and_patch(component_row, note_row):
    component = str(component_row['Component']).strip()
    release = component_row['Release']
    software_component = str(note_row['Software Component']).strip().lower() if pd.notna(note_row['Software Component']) else ""
    software_component_version = note_row.get('Software Component Version', None)
    
    if component.lower() not in software_component.split(","):
        return False
    
    try:
        if component in ["KRNL64UC", "KERNEL", "KRNL64NUC"]:
            if pd.notna(software_component_version):
                release_version = convert_version_format(release) if pd.notna(release) else None
                component_version = convert_version_format(software_component_version)
                if release_version is not None and component_version is not None:
                    if release_version == component_version:
                        return True
        else:
            from_version = convert_version_format(note_row['From'])
            to_version = convert_version_format(note_row['To'])
            release_version = convert_version_format(release) if pd.notna(release) else None
            
            if from_version is not None and to_version is not None and release_version is not None:
                if from_version <= release_version <= to_version:
                    return True
    except ValueError:
        return False
    
    return False

def clean_impacted_notes(ws_red_notes):
        """
        Pulisce la colonna 'Note Number' rimuovendo i valori duplicati consecutivi.
        """
        note_number_col = None
        for col in range(1, ws_red_notes.max_column + 1):
            if ws_red_notes.cell(row=1, column=col).value == "Note Number":
                note_number_col = col
                break
        if note_number_col:
            previous_note = None
            for row in range(2, ws_red_notes.max_row + 1):  # Evita l'intestazione
                current_note = ws_red_notes.cell(row=row, column=note_number_col).value
                if current_note == previous_note:
                    ws_red_notes.cell(row=row, column=note_number_col, value="")
                else:
                    previous_note = current_note



