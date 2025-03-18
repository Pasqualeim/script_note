import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import re
import tkinter as tk
from tkinter import filedialog

import sys
import os
# Forza l'uso di UTF-8 su Windows
os.system("chcp 65001 > nul")  # Imposta la console di Windows su UTF-8
# Imposta stdout per non essere bufferizzato (stampa in tempo reale)
sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)


import sys
sys.path.append("/Users/pasqualeercolino/Documents/Developer/script_note/")  # Modifica con il tuo percorso

from def_mac import select_save_location, extract_sp_level, check_release_and_patch, clean_impacted_notes, convert_version_format, select_file_component, select_file_notes

def apply_color_to_note_number(components_df, notes_df, notes_file):
    output_filename = select_save_location("Note Extraction_Updated.xlsx")
    red_notes_filename = select_save_location("Impacted_Notes.xlsx")
    
    wb = load_workbook(notes_file)
    ws = wb.active  
    
    wb_red_notes = Workbook()
    ws_red_notes = wb_red_notes.active
    ws_red_notes.append(["Note Number", "Impacted Component", "From", "To", "Patch Level", "SPLevel"])
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for index, note_row in notes_df.iterrows():
        found = False
        impacted_components = []
        patch_level = extract_sp_level(note_row['Patch Level']) if 'Patch Level' in notes_df.columns and pd.notna(note_row['Patch Level']) else None
        sp_level_component = None
        
        for _, component_row in components_df.iterrows():
            if check_release_and_patch(component_row, note_row):
                sp_level = extract_sp_level(component_row['SPLevel']) if 'SPLevel' in components_df.columns and pd.notna(component_row['SPLevel']) else None
                
                print(f"Controllo componente: {component_row['Component']} | SPLevel: {sp_level}")
                
                if patch_level is None:
                    found = True
                    impacted_components.append(component_row['Component'])
                    sp_level_component = sp_level
                elif sp_level is not None and patch_level > sp_level:
                    found = True
                    impacted_components.append(component_row['Component'])
                    sp_level_component = sp_level
        
        if found:
            note_row_idx = index + 2  
            note_cell = ws[f"A{note_row_idx}"]  
            
            for merged_range in ws.merged_cells.ranges:
                if note_cell.coordinate in merged_range:
                    note_cell = ws[merged_range.start_cell.coordinate]
                    break
            
            note_cell.fill = red_fill
            print(f"Impattato: {note_cell.coordinate} colorato di rosso")
            
            ws_red_notes.append([
                note_cell.value,
                ", ".join(impacted_components),
                note_row['From'],
                note_row['To'],
                patch_level if 'Patch Level' in notes_df.columns else None,
                sp_level_component
            ])
    
    wb.save(output_filename)
    clean_impacted_notes(ws_red_notes)
    wb_red_notes.save(red_notes_filename)
    print(f"✅ Salvataggio completato")

def main():
    components_file = select_file_component("Seleziona il file delle componenti.")
    if not components_file:
        print("Nessun file selezionato per le componenti.")
        return

    notes_file = select_file_notes("Seleziona il file Note Extraction.xlsx")
    if not notes_file:
        print("Nessun file selezionato per le note.")
        return

    components_df = pd.read_excel(components_file)
    notes_df = pd.read_excel(notes_file)
    
    apply_color_to_note_number(components_df, notes_df, notes_file)

# ⚠ Evita l'esecuzione automatica quando importato in GUI.py
if __name__ == "__main__":
    main()