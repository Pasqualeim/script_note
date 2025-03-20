import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import re

def convert_version_format(version):
    """
    Converte una versione con punto (es. 7.22) in un numero intero (es. 722) per l'analisi.
    """
    try:
        return int(str(version).replace(".", "")) if pd.notna(version) else None
    except ValueError:
        return None

def extract_sp_level(sp_value):
    """
    Estrae il numero SP (es. SP007 -> 7) per il confronto con SP-Level.
    """
    match = re.search(r"SP(\d+)", str(sp_value))
    return int(match.group(1).lstrip("0")) if match and match.group(1).lstrip("0") else None

def extract_patch_level(patch_value):
    """
    Estrae il Patch Level dalla stringa, ignorando zeri iniziali e finali e le lettere.
    """
    if pd.notna(patch_value):
        patch_str = re.sub(r'[^0-9]', '', str(patch_value))  # Mantieni solo i numeri
        return int(patch_str) if patch_str else None
    return None

def normalize_version(version):
    """
    Rimuove i caratteri non numerici e gli zeri iniziali e finali da una versione.
    """
    if pd.notna(version):
        version = re.sub(r'[^0-9]', '', str(version))
        return version.lstrip('0').rstrip('0')  # Rimuove gli zeri iniziali e finali
    return None

def check_release_and_patch(component_row, note_row):
    component = str(component_row['Component']).strip()
    release = component_row['Release']  # Manteniamo il formato originale
    software_component = str(note_row['Software Component']).strip() if pd.notna(note_row['Software Component']) else ""

    if component not in software_component:
        return False

    try:
        from_version = convert_version_format(note_row['From'])
        to_version = convert_version_format(note_row['To'])
        release_version = convert_version_format(release) if pd.notna(release) else None

        if from_version is not None and to_version is not None and release_version is not None:
            if from_version <= release_version <= to_version:
                return True
    except ValueError:
        return False
    return False

def apply_color_to_note_number(components_df, notes_df, output_filename="Note Extraction_Updated.xlsx", red_notes_filename="Impacted_Notes.xlsx"):
    wb = load_workbook("Note Extraction.xlsx")
    ws = wb.active  

    # Creazione di un nuovo file Excel per le sole note impattate
    wb_red_notes = Workbook()
    ws_red_notes = wb_red_notes.active
    ws_red_notes.append(list(notes_df.columns))  # Aggiunta delle intestazioni

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    print(f"Colonne nel DataFrame 'notes_df': {notes_df.columns}")

    for index, note_row in notes_df.iterrows():
        found = False
        impacted_components = []  # 🔹 Aggiunto per registrare le componenti impattate
        patch_level = extract_patch_level(note_row['Patch Level']) if 'Patch Level' in notes_df.columns and pd.notna(note_row['Patch Level']) else None

        if patch_level is not None:
            print(f"Patch Level trovato per la nota {note_row['Note Number']}: {patch_level}")

        for _, component_row in components_df.iterrows():
            if check_release_and_patch(component_row, note_row):
                software_version = normalize_version(note_row['Software Component Version'])
                release_version = normalize_version(component_row['Release'])

                print(f"Verifica versione: software_version = {software_version}, release_version = {release_version}")

                if software_version is not None and release_version is not None:
                    if software_version == release_version:
                        sp_level = extract_sp_level(component_row['SP-Level']) if 'SP-Level' in components_df.columns and pd.notna(component_row['SP-Level']) else None
                        print(f"SP-Level della componente: {sp_level}")

                        # Confronto tra Patch Level e SP-Level
                        if patch_level is not None and sp_level is not None:
                            if patch_level > sp_level:  # Se Patch Level della nota è maggiore dello SP-Level della componente
                                found = True
                                impacted_components.append(component_row['Component'])  # 🔹 Registra la componente impattata
                                print(f"🔴 Patch Level ({patch_level}) > SP-Level ({sp_level}). Nota colorata in rosso.")
                                break  # 🔹 Esce dal loop dopo aver trovato un match impattato

        if found:
            print(f"Nota {note_row['Note Number']} è impattata.")
        else:
            print(f"Nota {note_row['Note Number']} NON è impattata.")

        if found:
            note_row_idx = index + 2  # Riga della nota da colorare (considerando l'intestazione)
            note_cell = ws[f"A{note_row_idx}"]  # Cella da colorare in rosso

            for merged_range in ws.merged_cells.ranges:
                if note_cell.coordinate in merged_range:
                    print(f"Cella {note_cell.coordinate} è unita. Inizio dell'intervallo unito: {merged_range.start_cell.coordinate}")
                    note_cell = ws[merged_range.start_cell.coordinate]  # Prendi la cella principale
                    break

            note_cell.fill = red_fill
            print(f"Impattato: {note_cell.coordinate} colorato di rosso")

            merged_row_values = [note_cell.value] + impacted_components + [
                note_row['From'], note_row['To'], patch_level, sp_level
            ]

            ws_red_notes.append(merged_row_values)

    wb.save(output_filename)
    wb_red_notes.save(red_notes_filename)
    print(f"✅ Salvataggio completato: {output_filename} e {red_notes_filename}")

# Lettura dei dati dai file Excel
components_df = pd.read_excel("Components.xlsx")
notes_df = pd.read_excel("Note Extraction.xlsx")

# Chiamata alla funzione di elaborazione
apply_color_to_note_number(components_df, notes_df)
