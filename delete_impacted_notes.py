from openpyxl import load_workbook

# Carica il file delle note impattate
file_path = "Impacted_Notes.xlsx"
wb = load_workbook(file_path)
ws = wb.active

# Trova la colonna della "Note Number"
note_number_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Note Number":
        note_number_col = col
        break

# Se la colonna Note Number esiste, processa le righe
if note_number_col:
    previous_note = None
    for row in range(2, ws.max_row + 1):  # Evita l'intestazione
        current_note = ws.cell(row=row, column=note_number_col).value

        # Se la nota è uguale alla precedente, cancella il valore dalla cella (ma non la riga intera)
        if current_note == previous_note:
            ws.cell(row=row, column=note_number_col, value="")  # Imposta la cella come vuota
        else:
            previous_note = current_note  # Aggiorna il valore precedente

# Salva il file aggiornato
updated_file_path = "Impacted_Notes_Cleaned.xlsx"
wb.save(updated_file_path)

print(f"✅ File aggiornato salvato come: {updated_file_path}")