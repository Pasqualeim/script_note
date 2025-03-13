import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import re
import tkinter as tk
from tkinter import filedialog

def select_file(prompt):
    """
    Mostra una finestra di dialogo per selezionare un file e restituisce il percorso.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    return file_path

def select_save_location(default_name):
    """
    Mostra una finestra di dialogo per selezionare la posizione di salvataggio del file.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        title="Seleziona dove salvare il file",
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    return file_path

def convert_version_format(version):
    """
    Converte una versione con punto (es. 7.22) in un numero intero (es. 722) per l'analisi.
    """
    try:
        return int(str(version).replace(".", "")) if pd.notna(version) else None
    except ValueError:
        return None

def extract_sp_level(sp_value):
    """
    Legge il valore numerico di SPLevel direttamente senza regex, poiché il formato è numerico.
    """
    try:
        return int(sp_value) if pd.notna(sp_value) else None
    except ValueError:
        return None

def check_release_and_patch(component_row, note_row):
    component = str(component_row['Component']).strip()
    release = component_row['Release']
    software_component = str(note_row['Software Component']).strip().lower() if pd.notna(note_row['Software Component']) else ""
    
    if component.lower() not in software_component.split(","):
        return False
    
    try:
        from_version = convert_version_format(note_row['From'])
        to_version = convert_version_format(note_row['To'])
        release_version = convert_version_format(release) if pd.notna(release) else None
        
        if from_version is not None and to_version is not None and release_version is not None:
            if from_version <= release_version <= to_version:
                return True
    except ValueError:
        return False
    
    return False

def apply_color_to_note_number(components_df, notes_df, notes_file):
    output_filename = select_save_location("Note Extraction_Updated.xlsx")
    red_notes_filename = select_save_location("Impacted_Notes.xlsx")
    
    wb = load_workbook(notes_file)
    ws = wb.active  
    
    wb_red_notes = Workbook()
    ws_red_notes = wb_red_notes.active
    ws_red_notes.append(["Note Number", "Impacted Component", "From", "To", "Patch Level", "SPLevel"])
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for index, note_row in notes_df.iterrows():
        found = False
        impacted_components = []
        patch_level = extract_sp_level(note_row['Patch Level']) if 'Patch Level' in notes_df.columns and pd.notna(note_row['Patch Level']) else None
        sp_level_component = None
        
        for _, component_row in components_df.iterrows():
            if check_release_and_patch(component_row, note_row):
                sp_level = extract_sp_level(component_row['SPLevel']) if 'SPLevel' in components_df.columns and pd.notna(component_row['SPLevel']) else None
                
                print(f"📌 Controllo componente: {component_row['Component']} | SPLevel: {sp_level}")
                
                if patch_level is None:
                    found = True
                    impacted_components.append(component_row['Component'])
                    sp_level_component = sp_level
                elif sp_level is not None and patch_level > sp_level:
                    found = True
                    impacted_components.append(component_row['Component'])
                    sp_level_component = sp_level
        
        if found:
            note_row_idx = index + 2  
            note_cell = ws[f"A{note_row_idx}"]  
            
            for merged_range in ws.merged_cells.ranges:
                if note_cell.coordinate in merged_range:
                    note_cell = ws[merged_range.start_cell.coordinate]
                    break
            
            note_cell.fill = red_fill
            print(f"🔴 Impattato: {note_cell.coordinate} colorato di rosso")
            
            ws_red_notes.append([
                note_cell.value,
                ", ".join(impacted_components),
                note_row['From'],
                note_row['To'],
                patch_level if 'Patch Level' in notes_df.columns else None,
                sp_level_component
            ])
    
    wb.save(output_filename)
    wb_red_notes.save(red_notes_filename)
    print(f"✅ Salvataggio completato: {output_filename} e {red_notes_filename}")

components_file = filedialog.askopenfilename(title="Seleziona il file Components.xlsx", filetypes=[("Excel files", "*.xlsx")])
components_df = pd.read_excel(components_file)
notes_file = filedialog.askopenfilename(title="Seleziona il file Note Extraction.xlsx", filetypes=[("Excel files", "*.xlsx")])
notes_df = pd.read_excel(notes_file)
apply_color_to_note_number(components_df, notes_df, notes_file)
